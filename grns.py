import os, shutil, openpyxl, PyPDF2, re, zlib, fnmatch
import time
import pprint
import openpyxl
import pprint
from win32com import client
from pathlib import Path


# sql code to generation table (grnxl.xlsx)
# SELECT  purchase_order_line.order_no ,purchase_order_line.requisition_no, purchase_order_line.line_no, purchase_order_line.description,purchase_order_line.buy_unit_meas, purchase_order_line.original_qty,purchase_order_line.original_qty ,purchase_order_line.part_no, receipt_info.delivery_date,supplier.vendor_name FROM  purchase_order_line
# JOIN receipt_info
# ON purchase_order_line.order_no = receipt_info.source_ref1
# JOIN purchase_order
# ON purchase_order.order_no = purchase_order_line.order_no
# JOIN supplier
# ON purchase_order.vendor_no = supplier.vendor_no
# WHERE purchase_order_line.state =  'Closed' and purchase_order_line.order_no IN ('9981','11060','11427','11649','11686','12052','12103','12111','12144','12161','12208','12230','12233','12303','12370','12385','12396','12422','12423','12468','12492','12520','12532','12553','12646','12782','12837','13162','12217','12371','13082','13118')


grntemplates = openpyxl.load_workbook("grnxl.xlsx")
active_sheet = grntemplates.active
data = {}

for row in range(2, active_sheet.max_row + 1):
    PO = active_sheet["A" + str(row)].value
    PR = active_sheet["B" + str(row)].value
    UOM = active_sheet["E" + str(row)].value
    line_no = active_sheet["C" + str(row)].value
    part_des = active_sheet["D" + str(row)].value
    ord_qty = active_sheet["F" + str(row)].value
    del_qty = active_sheet["G" + str(row)].value
    ifs_num = active_sheet["H" + str(row)].value
    date = active_sheet["I" + str(row)].value
    supplier = active_sheet["J" + str(row)].value
    line_details = (
        str(line_no)
        + ";;;"
        + str(part_des)
        + ";;;"
        + str(UOM)
        + ";;;"
        + str(ord_qty)
        + ";;;"
        + str(del_qty)
        + ";;;"
        + str(ifs_num)
    )

    data.setdefault(
        PO, {"PO": PO, "PR": PR, "date": date, "supplier": supplier, "line_details": {}}
    )
    data[PO]["line_details"].setdefault(line_no, line_details)


print(pprint.pformat(data))


for po in data.keys():
    grn = openpyxl.load_workbook("GRN Template.xlsx")
    sheet = grn.active
    sheet["C8"].value = data[po]["supplier"]
    sheet["C11"].value = po
    sheet["C12"].value = data[po]["PR"]
    sheet["F8"].value = data[po]["date"]
    id = 15
    ln_no = 1
    for i in range(1, len(data[po]["line_details"]) + 1):
        try:
            line_item = data[po]["line_details"][i]
            list = line_item.split(";;;")
            if list[1] == " Lots " or list[2] == "Lot" or list[2] == "Sheet":
                continue

            sheet["B" + str(id)].value = ln_no
            sheet["C" + str(id)].value = list[1]
            sheet["D" + str(id)].value = list[2]
            sheet["E" + str(id)].value = list[3]
            sheet["F" + str(id)].value = list[4]
            if list[5] == "None":
                continue
            else:
                sheet["G" + str(id)].value = list[5]
            id += 1
            ln_no += 1
        except:
            continue

    excel_filename = str(po) + ".xlsx"
    grn.save(excel_filename)
    pdf_filename = str(po) + ".pdf"
    input_path = os.path.abspath(Path(excel_filename))
    output_path = os.path.abspath(Path(pdf_filename))
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False
    Workbook = app.Workbooks.Open(input_path)
    try:
        Workbook.ActiveSheet.ExportAsFixedFormat(0, output_path)
    except Exception as e:
        print(
            "Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again"
        )
        print(str(e))
    Workbook.Close()

print("Working on file,pr and manifest now")

pdf_file = []


for file in os.listdir(os.getcwd()):
    if file.endswith(".pdf"):
        pdf_file.append(file)
    else:
        continue

# Creating folder for each PO with folder name as the PO name
PR_found = []
for files in pdf_file:
    if files == 'None':
        continue
    pos_dir = Path(os.getcwd()) / files
    excel_files = Path(os.getcwd()) / str(files[: len(files) - 4]+ '.xlsx')
    new_folder = Path(os.getcwd()) / str(files[: len(files) - 4])
    os.makedirs(new_folder, exist_ok=True)
    if os.path.isfile(new_folder / str(files[: len(files) - 4]+ '.xlsx')):
        os.remove(new_folder / str(files[: len(files) - 4]+ '.xlsx'))
    elif os.path.isdir(new_folder / str(files[: len(files) - 4]+ '.xlsx')):
        shutil.rmtree(new_folder / str(files[: len(files) - 4]+ '.xlsx'))

    if os.path.isfile(new_folder / files):
        os.remove(new_folder / files)
    elif os.path.isdir(new_folder / files):
        shutil.rmtree(new_folder / files)
    shutil.move(excel_files, new_folder)
    shutil.move(pos_dir, new_folder)

    # finding their PO's respective PR's
    excel_sheet = openpyxl.load_workbook("grnxl.xlsx")
    sheet = excel_sheet.active
    for row in range(2, sheet.max_row + 1):
        try:
            if sheet.cell(row=row, column=1).value == int(files[: len(files) - 4]):
                PR_NO = sheet.cell(row=row, column=2).value
                PR_dir = Path(os.getcwd()) / "PR"
                for folder, subfolder, files in os.walk(PR_dir):
                    for file in files:
                        if re.findall(f".*{PR_NO}.*", str(file)):
                            shutil.copy(os.path.join(folder, file), new_folder)
                            PR_found.append(f"PO: {files[: len(files) - 4]}")
                        else:
                            continue

            else:
                continue
        except:
            continue
    try:
        folder_names = re.compile(r"^[0-9]+$")
        folders = []
        for file in os.listdir(os.getcwd()):
            if folder_names.search(file):
                folders.append(file)
            else:
                continue

        for folder in folders:
            folder_path = Path(os.getcwd()) / folder
            if len(os.listdir(folder_path)) < 2:
                pass
            else:
                continue
    except:
        continue

print(pdf_file)


# finding signed manifest name
manifest_path = Path(r"C:\Users\104535brbo\Desktop\Manifest")
signed_manifest_path = Path(r"C:\Users\104535brbo\Desktop\Signed manifests")
manifest_list = []

for folders, subfolders, files in os.walk(manifest_path):
    for file in files:
        if file.endswith(".pdf"):
            manifest_list.append(os.path.join(folders, file))
        else:
            continue


po_strings = []
for file in pdf_file:
    po_strings.append(str("PO: " + file[: len(file) - 4]))
print(po_strings)


for po in po_strings:
    found_manifest = []

    for file_dir in manifest_list:
        try:
            pdf_file = open(file_dir, "rb")
            pdf_reader = PyPDF2.PdfFileReader(pdf_file, strict=False)
            total_page = pdf_reader.numPages
            for page in range(0, total_page):
                pdf_getpage = pdf_reader.getPage(page)
                pdf_extract = pdf_getpage.extractText()
                if po in pdf_extract:
                    found_manifest.append(file_dir)
                else:
                    continue
        except (
            PyPDF2.utils.PdfReadError,
            KeyError,
            ValueError,
            TypeError,
            NameError,
            zlib.error,
        ):

            pass
    print(f"{po} : {found_manifest}")
    not_found_manifest = []
    # finding signed manifest using the manifest name found from above
    try:
        search_containing = []
        for manifest in found_manifest:
            search_containing.append(Path(manifest).stem)
        print(f"{po} :{search_containing}")
        signed_manifest_found = []

        for folders, subfolders, files in os.walk(signed_manifest_path):
            for file in files:
                file_path = os.path.join(folders, file)
                key_search_containing = []
                for path in search_containing:
                    key_search_containing = path[:34]
                    aa = key_search_containing[:12] + "0" + key_search_containing[12:33]
                    index = re.findall(f".*{aa}.*", str(file_path))
                    try:
                        if (
                            fnmatch.fnmatch(file_path, f"*{key_search_containing}*")
                            or re.findall(f".*{aa}.*", str(file_path))
                            or re.findall(f".*{aa[:18]}.*", str(file_path))
                        ):
                            shutil.copy(
                                file_path,
                                Path(r"C:\Users\104535brbo\Desktop\new dossiers")
                                / po[4:],
                            )
                        else:
                            continue
                    except:
                        continue

    except ValueError:
        pass

os.makedirs("no_manifest_and_pr", exist_ok=True)
os.makedirs("no_manifest", exist_ok=True)
os.makedirs("no_pr", exist_ok=True)
os.makedirs("complete_pos", exist_ok=True)


for folder, subfolder, file in os.walk(Path.cwd()):
    pos = re.compile(r"(\d{4,5})")
    if not pos.search(Path(folder).name):
        continue
    else:
        print(folder)
        pr = re.compile(r"^PR")
        manifest = re.compile(r"^(\d{3}\.)")
        grn = re.compile(r"^(\d{4,5})")
        files = os.listdir(os.path.join(Path.cwd(), folder))
        pr_status = False
        manifest_status = False
        grn_status = False
        for name in files:
            if pr.search(name):
                pr_status = True
            elif manifest.search(name):
                manifest_status = True
            elif grn.search(name):
                grn_status = True
        try:
            if pr_status == True and manifest_status == True and grn_status == True:
                shutil.move(folder, Path.cwd() / "complete_pos" / Path(folder).name)
            elif manifest_status == True and grn_status == True and pr_status == False:
                shutil.move(folder, Path.cwd() / "no_pr" / Path(folder).name)
            elif manifest_status == False and grn_status == True and pr_status == True:
                shutil.move(folder, Path.cwd() / "no_manifest" / Path(folder).name)
            elif manifest_status == False and pr_status == False:
                shutil.move(
                    folder, Path.cwd() / "no_manifest_and_pr" / Path(folder).name
                )
        except:
            continue
