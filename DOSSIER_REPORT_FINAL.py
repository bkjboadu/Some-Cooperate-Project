import os, shutil, openpyxl, PyPDF2, re, zlib, fnmatch
import pprint
from win32com import client
from openpyxl.styles import Alignment
from pathlib import Path
from collections import namedtuple
import pandas as pd

class Dossier:
    def __init__(self,grn_filename):
        self.grn_filename = grn_filename
        self.data = {}
        self.pdf_file = []

    def process_dossier(self):
        # self.convert_datetime()
        self.pos_dict()
        # self.create_grn()
        # self.locate_group_PR()
        self.locate_signedManifest()
        self.Group_file()

    def convert_datetime(self):
        x = pd.read_excel('grnxl.xlsx')
        x['Delivery_Date'] = pd.to_datetime(x['Delivery_Date']).dt.strftime("%d/%b/%y")
        x.to_excel('grnxl.xlsx', index=False)

    def pos_dict(self):
        active_sheet = openpyxl.load_workbook(self.grn_filename).active
        self.data = {}

        '''Extract individual details for each unique PO into a dictionary'''
        for row in range(2, active_sheet.max_row + 1):
            Details = namedtuple('Details', "PO PR UOM line_no part_des ord_qty del_qty ifs_num date supplier")
            details = Details(active_sheet["A" + str(row)].value, active_sheet["B" + str(row)].value,
                              active_sheet["E" + str(row)].value, active_sheet["C" + str(row)].value,
                              active_sheet["D" + str(row)].value,
                              active_sheet["F" + str(row)].value, active_sheet["G" + str(row)].value,
                              active_sheet["H" + str(row)].value, active_sheet["I" + str(row)].value,
                              active_sheet["J" + str(row)].value)
            line_details = (";;;".join(
                [str(details.line_no), str(details.part_des), str(details.UOM), str(details.ord_qty),
                 str(details.del_qty), str(details.ifs_num)]))
            self.data.setdefault(details.PO, {"PO": details.PO, "PR": details.PR, "date": details.date,
                                              "supplier": details.supplier, "line_details": {}})
            self.data[details.PO]["line_details"].setdefault(details.line_no, line_details)

        print(pprint.pformat(self.data))
        # return self.data


    def create_grn(self):

        '''sql code to extract self.data from ifs, we need to change the PO's in the last line of this code to what we will be working on currently'''

        # SELECT  purchase_order_line.order_no ,purchase_order_line.requisition_no, purchase_order_line.line_no, purchase_order_line.description,purchase_order_line.buy_unit_meas, purchase_order_line.original_qty,purchase_order_line.original_qty ,purchase_order_line.part_no, receipt_info.delivery_date,supplier.vendor_name FROM  purchase_order_line
        # JOIN receipt_info
        # ON purchase_order_line.order_no = receipt_info.source_ref1
        # JOIN purchase_order
        # ON purchase_order.order_no = purchase_order_line.order_no
        # JOIN supplier
        # ON purchase_order.vendor_no = supplier.vendor_no
        # WHERE purchase_order_line.state =  'Closed' and purchase_order_line.order_no IN ('11727','11847','11902','12100','12279','12370','12564','12755','13168','7613','13396','13401','13417','14395','14544','13649','14042','14119','14195','14350','15410','14440','14450','14504','14514','14517','12070','12245','13387','13522','13558','13821','14850','14028','14368','14589','15097','14663','14677','14702','14757','14927','14933','14999','14698','14842','14866','15096')
        #
        # "This function creates the grn based on the excel sheet provided."
        active_sheet = openpyxl.load_workbook(self.grn_filename).active
        self.data = {}

        '''Extract individual details for each unique PO into a dictionary'''
        for row in range(2, active_sheet.max_row + 1):
            Details = namedtuple('Details', "PO PR UOM line_no part_des ord_qty del_qty ifs_num date supplier")
            details = Details(active_sheet["A" + str(row)].value, active_sheet["B" + str(row)].value,
                              active_sheet["E" + str(row)].value, active_sheet["C" + str(row)].value,
                              active_sheet["D" + str(row)].value,
                              active_sheet["F" + str(row)].value, active_sheet["G" + str(row)].value,
                              active_sheet["H" + str(row)].value, active_sheet["I" + str(row)].value,
                              active_sheet["J" + str(row)].value)
            line_details = (";;;".join(
                [str(details.line_no), str(details.part_des), str(details.UOM), str(details.ord_qty),
                 str(details.del_qty), str(details.ifs_num)]))
            self.data.setdefault(details.PO, {"PO": details.PO, "PR": details.PR, "date": details.date,
                                              "supplier": details.supplier, "line_details": {}})
            self.data[details.PO]["line_details"].setdefault(details.line_no, line_details)

        print(pprint.pformat(self.data))

        '''from the dictionary created in the last block of code before this, we will fill in the grn templatewith this self.data accordingly'''
        for po in self.data.keys():
            grn = openpyxl.load_workbook("GRN Template.xlsx")
            sheet = grn.active
            sheet["C8"].value = self.data[po]["supplier"]
            sheet["C11"].value = po
            sheet["C12"].value = self.data[po]["PR"]
            sheet["F8"].value = self.data[po]["date"]
            id = 15
            ln_no = 1
            for i in range(1, len(self.data[po]["line_details"]) + 1):
                try:
                    line_item = self.data[po]["line_details"][i]
                    list = line_item.split(";;;")
                    if list[1] == " Lots " or list[2] == "Lot" or list[2] == "Sheet":
                        continue

                    sheet["B" + str(id)].value = ln_no
                    sheet["C" + str(id)].value = list[1]
                    sheet["D" + str(id)].value = list[2]
                    sheet["E" + str(id)].value = list[3]
                    sheet["F" + str(id)].value = list[4]
                    if list[5] == "None":
                        sheet["G" + str(id)].value = ' '
                    else:
                        sheet["G" + str(id)].value = list[5]

                    sheet["B" + str(id)].alignment = Alignment(wrap_text=True,vertical='center')
                    sheet["C" + str(id)].alignment = Alignment(wrap_text=True,vertical='center')
                    sheet["D" + str(id)].alignment = Alignment(wrap_text=True,vertical='center')
                    sheet["E" + str(id)].alignment = Alignment(wrap_text=True,vertical='center')
                    sheet["F" + str(id)].alignment = Alignment(wrap_text=True,vertical='center')
                    sheet["G" + str(id)].alignment = Alignment(wrap_text=True,vertical='center')

                    id += 1
                    ln_no += 1
                except:
                    continue



            for row in active_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            #adding signature and eni logo
            sheet['C15'].alignment = Alignment(wrap_text=True, vertical='center')
            signature = openpyxl.drawing.image.Image('signature.png')
            signature.anchor = 'C36'

            eni = openpyxl.drawing.image.Image('eni.png')
            eni.anchor = 'B2'

            sheet['H1'].alignment = Alignment(wrap_text=True, vertical='center')

            sheet.add_image(signature)
            sheet.add_image(eni)

            excel_filename = str(po) + ".xlsx"
            if excel_filename == 'None.xlsx':
                continue
            grn.save(excel_filename)
            pdf_filename = str(po) + ".pdf"
            input_path = os.path.abspath(Path(excel_filename))
            output_path = os.path.abspath(Path(pdf_filename))
            app = client.DispatchEx("Excel.Application")
            app.Interactive = False
            app.Visible = False
            Workbook = app.Workbooks.Open(input_path)
            try:
                Workbook.ActiveSheet.ExportAsFixedFormat(0, output_path)
            except Exception as e:
                print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
                print(str(e))
            Workbook.Close()

    def locate_group_PR(self):
        '''Finding the purchase requisition for each PO and moving it into the specific PO folder'''
        self.pdf_file = [file for file in Path.cwd().glob('*.pdf')]

        for files in self.pdf_file:
            NewFolder = os.path.splitext(files)[0]
            PurchaseOrder_excel = os.path.splitext(files)[0] + '.xlsx'
            os.makedirs(NewFolder,exist_ok=True)
            shutil.move(PurchaseOrder_excel,NewFolder)
            shutil.move(files,NewFolder,)

        for po in self.data.keys():
            NewFolder = os.path.join(Path.cwd(),str(po))
            PurchaseOrderRequisition = self.data[po]['PR']
            for prs in Path('PR').iterdir():
                if re.match(f".*{PurchaseOrderRequisition}.*",str(prs)):
                    shutil.copy(prs,NewFolder)

    def locate_signedManifest(self):
        '''Finding the signed manifest by first reading through the manifest since it is not scanned and then using the file names of the
        manifest found to find the signed manifest'''
        manifest_path = Path(r"C:\Users\104535brbo\Desktop\Manifest")
        signed_manifest_path = Path(r"C:\Users\104535brbo\Desktop\Signed manifests")

        '''Finding pdf in manifest paths'''
        for po in self.data.keys():
            print(po)
            found_manifest = []
            if po == None:
                continue
            for file in manifest_path.glob('*.pdf'):
                '''Opening files in manifest folder and check if they contain some PO numbers'''
                try:
                    with open(file, "rb") as pdf_file:
                        pdf_reader = PyPDF2.PdfFileReader(pdf_file,strict=False)
                        total_page = pdf_reader.numPages
                        pdf_extract = ''
                        for page in range(0, total_page):
                            pdf_getpage = pdf_reader.getPage(page)
                            pdf_extract += pdf_getpage.extractText()
                        if f"PO {str(po)}" in pdf_extract:
                            found_manifest.append(file)
                        if f"PO: {str(po)}" in pdf_extract:
                            found_manifest.append(file)
                        if f"PO:{str(po)}" in pdf_extract:
                            found_manifest.append(file)
                        else:
                            continue
                except (PyPDF2.utils.PdfReadError,KeyError,ValueError,TypeError,NameError,zlib.error,OSError):
                    pass

            print(f"{po} : {found_manifest}")
            '''Finding signed manifest using the manifest name'''
            try:
                if len(found_manifest) == 0:
                    continue
                search_containing =[Path(l).stem for l in found_manifest]
                signed_files = [Path(file) for file in Path(r'C:\Users\104535brbo\Desktop\Signed manifests').glob('*.pdf')]
                print(f"{po} :{search_containing}")
                for path in search_containing:
                    key_search_containing = path[:]
                    short_path = path[:12] + "0" + path[12:17]
                    print(short_path)
                    print(key_search_containing)
                    try:
                        for file in signed_files:
                            if (fnmatch.fnmatch(file, f"*{key_search_containing}*") or re.findall(f".*{short_path}.*",str(file)) or re.findall(f".*{short_path}.*", str(file))) or fnmatch.fnmatch(file, f"*{key_search_containing[:18]}*"):
                                shutil.copy(file, Path(r"C:\Users\104535brbo\Documents\python code\new dossiers") / str(po))
                                print('done')
                    except:
                        continue
            except ValueError:
                pass

    def Group_file(self):
        os.makedirs("no_manifest_and_pr", exist_ok=True)
        os.makedirs("no_manifest", exist_ok=True)
        os.makedirs("no_pr", exist_ok=True)
        os.makedirs("complete_pos", exist_ok=True)

        for folder, subfolder, file in os.walk(Path.cwd()):
            pos = re.compile(r"(\d{4,5})")
            if not pos.search(Path(folder).name):
                continue
            else:
                print(folder)
                pr = re.compile(r"^PR")
                manifest = re.compile(r"^(\d{3}\.)")
                grn = re.compile(r"^(\d{4,5})")
                files = os.listdir(os.path.join(Path.cwd(), folder))
                pr_status = False
                manifest_status = False
                grn_status = False
                for name in files:
                    if pr.search(name):
                        pr_status = True
                    elif manifest.search(name):
                        manifest_status = True
                    elif grn.search(name):
                        grn_status = True
                try:
                    if pr_status == True and manifest_status == True and grn_status == True:
                        shutil.move(folder, Path.cwd() / "complete_pos" / Path(folder).name)
                    elif manifest_status == True and grn_status == True and pr_status == False:
                        shutil.move(folder, Path.cwd() / "no_pr" / Path(folder).name)
                    elif manifest_status == False and grn_status == True and pr_status == True:
                        shutil.move(folder, Path.cwd() / "no_manifest" / Path(folder).name)
                    elif manifest_status == False and pr_status == False:
                        shutil.move(folder, Path.cwd() / "no_manifest_and_pr" / Path(folder).name)
                except:
                    continue

if __name__ == '__main__':
    Dossier('grnxl.xlsx').process_dossier()