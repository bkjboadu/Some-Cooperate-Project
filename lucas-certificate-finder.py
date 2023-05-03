import os
from pathlib import Path
import PyPDF2,shutil
import openpyxl
covid21 = Path(r'C:\Users\104535brbo\Desktop\covid python\2020')
# wb= openpyxl.load_workbook(r'C:\Users\104535brbo\Desktop\covid python\2021\01.02E.xlsx')
for folders,subfolders,files in os.walk(covid21):
    for file in files:
        try:
            if not file.endswith('.xlsx'):
                pass
            else:
                excel_name = os.path.join(folders,file)
                sheet = openpyxl.load_workbook(excel_name)
                wb = sheet.active
                names = []
                for row in range(1,wb.max_row + 1):
                    if wb.cell(row=row,column=3).value not in [None,' ','NAME']:
                        names.append(wb.cell(row=row,column=3).value.upper())
                report = Path(r'C:\Users\104535brbo\Desktop\covid python\certificates')


                for folder,subfolder,file in os.walk(report):
                    for file in file:
                        if Path(os.path.join(folders,file)).stem.lower().startswith('yinson'):
                            cert = Path(os.path.join(folder,file))
                            pdf_file = PyPDF2.PdfFileReader(open(cert,'rb'))
                            page1 = pdf_file.getPage(0)
                            text = page1.extractText()
                            percentage = 0
                            for name in names:
                                if name.strip() in text:
                                    percentage += 1
                            if percentage/len(names) * 100 >= 70:
                                location = Path(r'C:\Users\104535brbo\Desktop\covid python\found_certs 2020') / Path(excel_name).stem
                                os.makedirs(location,exist_ok=True)
                                print(excel_name,'-->',cert)
                                shutil.copy(excel_name,location)
                                shutil.copy(cert,location)
                                break
                            else:
                                continue
        except:
            continue





